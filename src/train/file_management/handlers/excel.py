from logging import getLogger
from pathlib import Path

import openpyxl

from .handler import Handler

logger = getLogger(__name__)


class ExcelHandler(Handler):
    @staticmethod
    def read_dict(file: Path) -> tuple[list[str], list[dict]]:
        """Can raise `RuntimeError`."""
        wb = openpyxl.load_workbook(file.as_posix(), read_only=True, data_only=True)
        sheet: openpyxl.worksheet.worksheet.Worksheet | None = wb.active  # type: ignore ()

        if sheet is None:
            msg = "Could not read excel sheet"
            raise RuntimeError(msg)

        col_count = sheet.max_column
        if col_count is None:
            logger.error(
                "ERROR! Openpyxl is not reporting correct column count, "
                "defaulting to 20",
            )
            col_count = 20

        headers = [str(sheet.cell(1, col + 1).value) for col in range(col_count)]
        data = [
            {
                headers[i]: str(row[i].value) if row[i].value is not None else ""
                for i in range(col_count)
            }
            for row in sheet.iter_rows(min_row=2, min_col=1, max_col=col_count)
        ]

        wb.close()
        return headers, data

    @staticmethod
    def write_dict(file: Path, headers: list[str], data: list[dict]) -> None:
        # print(data)
        wb = openpyxl.Workbook(write_only=True)
        sheet = wb.create_sheet()

        sheet.append(headers)
        for row in data:
            sheet.append(
                [row.get(heading, "") for heading in headers if heading is not None],
            )

        wb.save(file.as_posix())
        wb.close()
