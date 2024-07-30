from __future__ import annotations

import csv
import logging
from abc import ABC, abstractmethod
from datetime import time, timedelta
from enum import IntEnum, auto
from typing import TYPE_CHECKING, Any, TypeAlias

import openpyxl
import openpyxl.worksheet
import openpyxl.worksheet._write_only
import openpyxl.worksheet.worksheet
import openpyxl.writer.excel
from result import Err, Ok, Result

from train.db import decode_time, timediff
from train.exceptions import (
    InvalidFileDataError,
    InvalidHeadersError,
    UnsupportedFileTypeError,
)
from train.models.section import Section
from train.models.task import PartialTask, Task

if TYPE_CHECKING:
    from pathlib import Path
    from sqlite3 import Cursor

Reason: TypeAlias = str


class FileManager(ABC):
    class Format(IntEnum):
        bare_minimum = auto()
        mas_recent = auto()

    @staticmethod
    def _get_time(raw: str) -> time | None:
        try:
            return decode_time(raw)
        except ValueError:
            return None

    @staticmethod
    def try_int(raw: str, default: int = 0) -> int:
        try:
            return int(raw)
        except ValueError:
            return default

    @staticmethod
    def get_manager(
        src: Path,
        dst: Path,
        err: Path,
        logger: logging.Logger | None = None,
    ) -> FileManager:
        if logger is None:
            logger = logging.getLogger(__name__)

        if src.suffix == ".csv":
            return CSVManager(src, dst, err, logger)

        if dst.suffix == ".xlsx":
            return ExcelManager(src, dst, err, logger)

        raise UnsupportedFileTypeError(src.suffix)

    _headers: list[str]
    _fmt: Format | None

    def __init__(self, src: Path, dst: Path, err: Path, logger: logging.Logger) -> None:
        self.src_path = src
        self.dst_path = dst
        self.err_path = err
        self.logger = logger

    @property
    def headers(self) -> list[str]:
        return self._headers

    @property
    def fmt(self) -> Format:
        if self._fmt is None:
            raise ValueError

        return self._fmt

    @headers.setter
    def headers(self, headers: list[str]) -> None:
        self._headers = headers
        self._fmt = self.get_file_fmt_type(headers)

        is_valid = self._validate_headers()
        if not is_valid:
            raise InvalidHeadersError(self._fmt, self._headers)

    @staticmethod
    def get_file_fmt_type(headers: list[str]) -> Format:
        if "LOCATION" in headers:
            return FileManager.Format.mas_recent

        return FileManager.Format.bare_minimum

    def _validate_headers(self) -> bool:
        if self.fmt == FileManager.Format.bare_minimum:
            return {
                "date",  # Output
                "department",
                "den",
                "nature_of_work",
                "location",
                "block_section_or_yard",
                "corridor_block",
                "line",
                "demanded_time_from",
                "demanded_time_to",
                "block_demanded",
                "permitted_time_from",  # Output
                "permitted_time_to",  # Output
                "block_permitted",  # Output
            }.issubset(self._headers)

        if self.fmt == FileManager.Format.mas_recent:
            return {
                "DATE",  # Output
                "Department",
                "DEN",
                "Block Section/ Yard",
                "CORRIDOR block section",
                # "corridor block period",
                "UP/ DN Line",
                # "Block demanded in Hrs(Day or Night)",
                "Demanded time (From)",
                "Demanded time (To)",
                "Block demanded in(MINS)",
                "Permitted time (From) No need to fill",  # Output
                "Permitted time (To) No need to fill",  # Output
                "BLOCK PERMITTED MINS",  # Output
                # "Location - FROM",
                # "Location - TO",
                "Nautre of work & Quantum of Output Planned",
                # "Need for disconnection (If Yes Track Circuit and Signals Affected) Please give specific details without fail",  # noqa: E501
                # "Caution required",
                # "Caution required (if yes with relaxation date dd:mm:yyyy)",
                # "Power Block & its Elementary Section. Please give specific details without fail",  # noqa: E501
                # "Resources needed (M/C, Manpower, Supervisors) Include Crane,JCB,porcelain or any other equipment also",  # noqa: E501
                # "Whether site preparation & resources ready",
                # "Supervisors to be deputed (JE/SSE with section)",
                # "Coaching repercussions/ Movement Repercussions",
                # "Actual Block Granted From No need to fill",
                # "Actual Block Granted To No need to fill",
                # "Actual block duration MINS No need to fill",
                # "Over all % granted for the day No need to fill",
                # "Output as per Manual No need to fill",
                # "Actual Output",
                # "% Output vs Planned\nNo need to fill",
                # "% Output\nvs\nPlanned",
                # "PROGRESS",
                "LOCATION",
                # "SECTION",
                # "ARB/RB",
            }.issubset(self._headers)

        return True

    def encode_tasks(self, cur: Cursor, tasks: list[Task]) -> list[dict]:
        cur.execute(
            f"""
            SELECT
                DATE(task.starts_at),
                (
                    SELECT node.name FROM node
                    WHERE
                        node.id = section.from_id
                ),
                (
                    SELECT node.name FROM node
                    WHERE
                        node.id = section.to_id
                ),
                block.name,
                section.line,
                task.preferred_starts_at,
                task.preferred_ends_at,
                task.requested_duration / 60,
                TIME(task.starts_at),
                TIME(task.ends_at),
                task.priority,
                task.department,
                task.den,
                task.nature_of_work,
                task.location
            FROM task
            JOIN slot ON
                slot.id = task.slot_id
            JOIN section ON
                section.id = slot.section_id
            JOIN node ON
                node.id = section.from_id
            JOIN block ON
                block.id = node.block_id
            WHERE task.id IN ({', '.join(str(task.id) for task in tasks)})
            ORDER BY
                task.starts_at ASC
            """,  # noqa: S608
        )

        return [
            self.unmap(
                {
                    "priority": row[10],
                    "date": row[0],
                    "block_section_or_yard": (
                        f"{row[1].replace("_", " ")}"
                        if row[1] == row[2]
                        else f"{row[1]}-{row[2]}"
                    ),
                    "corridor_block": row[3],
                    "line": row[4],
                    "demanded_time_from": row[5],
                    "demanded_time_to": row[6],
                    "block_demanded": row[7],
                    "permitted_time_from": row[8],
                    "permitted_time_to": row[9],
                    "block_permitted": row[7],
                    "department": row[11],
                    "den": row[12],
                    "nature_of_work": row[13],
                    "location": row[14],
                },
            )
            for row in cur.fetchall()
        ]

    def map(self, item: dict) -> Result[dict[str, Any], Reason]:
        mapped = {}
        if self.fmt == FileManager.Format.bare_minimum:
            mapped = item | {"priority": 1}

        elif self.fmt == FileManager.Format.mas_recent:
            mapped = {
                "priority": 1,
                "date": item["DATE"],
                "block_section_or_yard": item["Block Section/ Yard"],
                "corridor_block": item["CORRIDOR block section"],
                "line": item["UP/ DN Line"],
                "demanded_time_from": item["Demanded time (From)"],
                "demanded_time_to": item["Demanded time (To)"],
                "block_demanded": item["Block demanded in(MINS)"],
                "permitted_time_from": item["Permitted time (From) No need to fill"],
                "permitted_time_to": item["Permitted time (To) No need to fill"],
                "block_permitted": item["BLOCK PERMITTED MINS"],
                "department": item["Department"],
                "den": item["DEN"],
                "nature_of_work": item["Nautre of work & Quantum of Output Planned"],
                "location": item["LOCATION"],
            }

        # Validate the type of data

        if not isinstance(mapped["priority"], int):
            self.logger.warning("Invalid priority")
            return Err("Invalid priority")

        if not isinstance(mapped["block_section_or_yard"], str):
            self.logger.warning("Invalid section")
            return Err("Invalid section")

        # if not isinstance(mapped["corridor_block"], str):
        #     self.logger.warning("Invalid block")
        #     return Err("Invalid block")

        if not isinstance(mapped["line"], str):
            self.logger.warning("Invalid line")
            return Err("Invalid line")

        mapped["demanded_time_from"] = FileManager._get_time(
            str(mapped["demanded_time_from"]).strip(),
        )

        mapped["demanded_time_to"] = FileManager._get_time(
            str(mapped["demanded_time_to"]).strip(),
        )

        mapped["block_demanded"] = FileManager.try_int(str(mapped["block_demanded"]), 0)

        return Ok(mapped)

    def unmap(self, item: dict) -> dict:
        if self.fmt == FileManager.Format.mas_recent:
            return {
                "DATE": item["date"],
                "Department": item["department"],
                "DEN": item["den"],
                "Block Section/ Yard": item["block_section_or_yard"],
                "CORRIDOR block section": item["corridor_block"],
                # # "corridor block period",
                "UP/ DN Line": item["line"],
                # # "Block demanded in Hrs(Day or Night)",
                "Demanded time (From)": item["demanded_time_from"],
                "Demanded time (To)": item["demanded_time_to"],
                "Block demanded in(MINS)": item["block_demanded"],
                "Permitted time (From) No need to fill": item["permitted_time_from"],
                "Permitted time (To) No need to fill": item["permitted_time_to"],
                "BLOCK PERMITTED MINS": item["block_permitted"],
                # # "Location - FROM",
                # # "Location - TO",
                "Nautre of work & Quantum of Output Planned": item["nature_of_work"],
                # # "Need for disconnection (If Yes Track Circuit and Signals Affected) Please give specific details without fail",  # noqa: E501
                # # "Caution required",
                # # "Caution required (if yes with relaxation date dd:mm:yyyy)",
                # # "Power Block & its Elementary Section. Please give specific details without fail",  # noqa: E501
                # # "Resources needed (M/C, Manpower, Supervisors) Include Crane,JCB,porcelain or any other equipment also",  # noqa: E501
                # # "Whether site preparation & resources ready",
                # # "Supervisors to be deputed (JE/SSE with section)",
                # # "Coaching repercussions/ Movement Repercussions",
                # "Actual Block Granted From No need to fill": item["actual_block_from"]
                # "Actual Block Granted To No need to fill": item["actual_block_to"],
                # "Actual block duration MINS No need to fill": item["block_actual"],
                # "Over all % granted for the day No need to fill": item["overall_per"],
                # "Output as per Manual No need to fill": item["output"],
                # "Actual Output",
                # "% Output vs Planned\nNo need to fill",
                # # "% Output\nvs\nPlanned",
                # # "PROGRESS",
                "LOCATION": item["location"],
                # "SECTION",
                # "ARB/RB",
            }

        # bare minimum
        return item

    def decode(self, cur: Cursor, item: dict) -> Result[PartialTask, Reason]:
        res = self.map(item)
        if isinstance(res, Err):
            return res

        mapped_item = res.value

        preferred_ends_at = mapped_item["demanded_time_to"]
        preferred_starts_at = mapped_item["demanded_time_from"]

        section = Section.find_by_name_and_line(
            cur,
            mapped_item["block_section_or_yard"],
            mapped_item["line"],
        )

        if section is None:
            self.logger.warning(
                "Could not find section: %s - %s",
                mapped_item["block_section_or_yard"],
                mapped_item["line"],
            )

            return Err(
                "Invalid section"
                f' {mapped_item["block_section_or_yard"]}-{mapped_item["line"]}',
            )

        if mapped_item["block_demanded"] == 0 and (
            preferred_starts_at is None or preferred_ends_at is None
        ):
            self.logger.warning("Invalid duration (It is empty)")
            return Err("Invalid duration (It is empty)")

        if (preferred_starts_at is None) ^ (preferred_ends_at is None):
            self.logger.warning("Demanded range is not complete")
            return Err("Demanded range is not complete")

        return Ok(
            PartialTask(
                mapped_item["priority"],
                (
                    timedelta(minutes=mapped_item["block_demanded"])
                    if mapped_item["block_demanded"] != 0
                    else timediff(preferred_starts_at, preferred_ends_at)
                ),
                preferred_starts_at,
                preferred_ends_at,
                section.id,
                mapped_item["department"],
                mapped_item["den"],
                mapped_item["nature_of_work"],
                mapped_item["location"],
            ),
        )

    def read(self, cur: Cursor) -> list[Result[PartialTask, Reason]]:
        data = self._read()

        taskqs = []
        for idx, item in enumerate(data):
            if isinstance(decoded := self.decode(cur, item), Err):
                self.logger.warning("Ignoring item on row `%d` ", idx)

            taskqs.append(decoded)

        return taskqs

    def write(self, cur: Cursor, tasks: list[Task]) -> None:
        data = self.encode_tasks(cur, tasks)
        self._write(data, self.headers, self.dst_path)

    def write_error(self, skipped_list: list[tuple[int, str]]):
        self._write(
            [{"row": row, "reason": reason} for row, reason in skipped_list],
            ["row", "reason"],
            self.err_path,
            with_color_coding=False,
        )

    @abstractmethod
    def _read(self) -> list[dict]: ...

    @abstractmethod
    def _write(
        self,
        data: list[dict],
        headers: list[str],
        file: Path,
        *,
        with_color_coding: bool = True,
    ) -> None: ...


class CSVManager(FileManager):
    def _read(self) -> list[dict]:
        with self.src_path.open(mode="r", newline="", encoding="utf-8-sig") as fd:
            reader = csv.DictReader(fd)
            data = [*reader]

        if not data:
            msg = "Please give file with data ;-;"
            raise InvalidFileDataError(msg)

        self.headers = [*data[0].keys()]
        return data

    def _write(
        self,
        data: list[dict],
        headers: list[str],
        file: Path,
        *,
        with_color_coding: bool = False,  # noqa: ARG002
    ) -> None:
        with file.open(mode="w", newline="") as fd:
            writer = csv.DictWriter(fd, headers)
            writer.writeheader()
            writer.writerows(data)


class ExcelManager(FileManager):
    def _read(self) -> list[dict]:
        wb = openpyxl.load_workbook(self.src_path, read_only=True, data_only=True)
        sheet: openpyxl.worksheet.worksheet.Worksheet | None = wb.active  # type: ignore ()

        if sheet is None:
            msg = "Could not read excel sheet"
            raise InvalidFileDataError(msg)

        col_count = sheet.max_column

        self.headers = [str(sheet.cell(1, col + 1).value) for col in range(col_count)]
        data = [
            {
                self.headers[i]: str(row[i].value) if row[i].value is not None else ""
                for i in range(col_count)
            }
            for row in sheet.iter_rows(min_row=2, min_col=1, max_col=col_count)
        ]

        wb.close()
        return data

    def _write(
        self,
        data: list[dict],
        headers: list[str],
        file: Path,
        *,
        with_color_coding: bool = True,
    ) -> None:
        wb = openpyxl.Workbook(write_only=True)
        sheet = wb.create_sheet()

        if with_color_coding:
            self._write_color_coded_headers(sheet)
        else:
            sheet.append(headers)

        for row in data:
            sheet.append([row.get(heading, "") for heading in headers])

        wb.save(file.as_posix())
        wb.close()

    def _write_color_coded_headers(
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
    ) -> None:
        from openpyxl.cell import Cell, WriteOnlyCell
        from openpyxl.styles import PatternFill

        if self.fmt == FileManager.Format.bare_minimum:
            ...

        elif self.fmt == FileManager.Format.mas_recent:
            OUTPUT_HEADERS = [  # noqa: N806
                "DATE",  # Output
                "Permitted time (From) No need to fill",  # Output
                "Permitted time (To) No need to fill",  # Output
                "BLOCK PERMITTED MINS",  # Output
            ]
            INPUT_HEADERS = [  # noqa: N806
                "Department",
                "DEN",
                "Block Section/ Yard",
                "CORRIDOR block section",
                "UP/ DN Line",
                "Demanded time (From)",
                "Demanded time (To)",
                "Block demanded in(MINS)",
                "Nautre of work & Quantum of Output Planned",
                "LOCATION",
            ]

            cols: list[Cell | None] = [None for i in range(len(self.headers))]
            for output_header in OUTPUT_HEADERS:

                cell = WriteOnlyCell(sheet, value=output_header)
                # cell.font = Font(name='Courier', size=36)
                cell.fill = PatternFill(start_color="FF00FF00", fill_type="solid")
                cols[self.headers.index(output_header)] = cell

            for input_header in INPUT_HEADERS:
                cell = WriteOnlyCell(sheet, value=input_header)
                # cell.font = Font(name='Courier', size=36)
                cell.fill = PatternFill(start_color="FFFFFF00", fill_type="solid")
                cols[self.headers.index(input_header)] = cell

            SCAM_HEADERS = (  # noqa: N806
                header
                for header in self.headers
                if header not in INPUT_HEADERS and header not in OUTPUT_HEADERS
            )
            for scam_header in SCAM_HEADERS:  # Robus application moment
                cell = WriteOnlyCell(sheet, value=scam_header)
                cell.fill = PatternFill(start_color="FFFF0000", fill_type="solid")
                cols[self.headers.index(scam_header)] = cell

            sheet.append(cols)
